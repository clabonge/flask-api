import openpyxl
import os
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("script_log.txt"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

def create_locker_configurator_excel(filename="LockerTowerConfigurator.xlsx"):
    logger.info(f"Starting Excel file creation: {filename}")
    try:
        # Create workbook
        logger.info("Creating new workbook")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Reference Sheet (Hidden)
        logger.info("Setting up Reference sheet")
        ref_sheet = wb.create_sheet("Reference")
        ref_data = [
            ("Parameter", "Value", "List"),
            ("Tower Heights (in)", "", "36,48,60,72"),
            ("Lock Controller Locations", "", "Top,Bottom,Both,None"),
            ("Control Housing", "", "Yes,No"),
            ("Materials", "", "MDF,Polywood,Polygood,M,P,G"),
            ("Material Costs ($/sqft)", "", "MDF=8,Polywood=10,Polygood=12"),
            ("Lock Types", "", "Basic,Medium,Premier,B,M,P"),
            ("Lock Costs ($)", "", "Basic=5,Medium=50,Premier=85"),
            ("Hinge Types", "", "Standard,Reinforced,S,R"),
            ("Hinge Costs ($)", "", "Standard=7.5,Reinforced=15"),
            ("Locker Heights (in)", "", "6,12,18,24,36,48,60,72"),
            ("Locker Widths (in)", "", "9,18"),
            ("Third Party Options", "", "Yes,No"),
            ("Width (in)", 18, ""),
            ("Depth (in)", 18, ""),
            ("Control Housing Height (in)", 18, ""),
            ("Labor Rate ($/hr)", 50, ""),
            ("CNC Cost ($)", 300, ""),
            ("Markup Percentage", 1.5, ""),
            ("Controller Heights (in)", "", "Top=4,Bottom=4,Both=8,None=0")
        ]
        for row, data in enumerate(ref_data, 1):
            for col, value in enumerate(data, 1):
                ref_sheet.cell(row=row, column=col).value = value
        ref_sheet.sheet_state = "hidden"

        # Project Configurator Sheet (First)
        logger.info("Setting up Project Configurator sheet")
        project_sheet = wb.create_sheet("Project Configurator", 0)
        project_sheet['A1'] = "Locker Tower Configurator - Project Setup"
        project_sheet['A1'].font = Font(bold=True, size=14)
        project_sheet['A2'] = "Enter the number of towers and controller panel locations below."
        project_sheet['A3'] = "Number of Towers"
        project_sheet['B3'] = 1
        project_sheet['A4'] = "Tower"
        project_sheet['B4'] = "Controller Panel"
        project_sheet['A4'].font = Font(bold=True)
        project_sheet['B4'].font = Font(bold=True)
        for col in ['A', 'B']:
            project_sheet[f'{col}4'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Pre-fill tower numbers and add dropdowns
        dv_towers = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
        dv_controller = DataValidation(type="list", formula1="Reference!C3", allow_blank=True)
        project_sheet.add_data_validation(dv_towers)
        project_sheet.add_data_validation(dv_controller)
        dv_towers.add("B3")
        for row in range(5, 15):  # Up to 10 towers
            project_sheet.cell(row=row, column=1).value = row - 4
            dv_controller.add(f"B{row}")

        project_sheet.column_dimensions['A'].width = 15
        project_sheet.column_dimensions['B'].width = 20

        # Summary Sheet (Placeholder, updated later)
        logger.info("Setting up Summary sheet")
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet['A1'] = "Locker Tower Configurator - Summary"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        summary_sheet['A2'] = "Results will appear here after entering tower details."
        summary_headers = [
            "Tower Number", "Base Tower Height (in)", "Adjusted Tower Height (in)",
            "Lock Controller Location", "Control Housing", "Material", "Lock Type",
            "Hinge Type", "Locker Configs", "Surface Area (sqft)", "Material Cost ($)",
            "Hinge Cost ($)", "Lock Cost ($)", "Hardware Cost ($)", "Labor Cost ($)",
            "CNC Cutting Cost ($)", "Total Cost ($)", "Sale Price ($)"
        ]
        for col, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Save the file
        logger.info(f"Saving to {os.path.abspath(filename)}")
        wb.save(filename)
        logger.info(f"Excel file saved successfully: {os.path.abspath(filename)}")

    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        raise

if __name__ == "__main__":
    try:
        create_locker_configurator_excel()
        print("Please open LockerTowerConfigurator.xlsx, enter the number of towers in the Project Configurator sheet, and save. Re-run this script to generate tower-specific sheets.")
    except Exception as e:
        logger.error(f"Script failed: {e}")
    input("Press Enter to exit...")