import openpyxl
import os
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("script_log.txt"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

def update_locker_configurator_excel(filename="LockerTowerConfigurator.xlsx"):
    logger.info(f"Updating Excel file: {filename}")
    try:
        # Load workbook
        logger.info("Loading workbook")
        wb = openpyxl.load_workbook(filename)

        # Get Project Configurator
        project_sheet = wb["Project Configurator"]
        num_towers = project_sheet['B3'].value
        if not isinstance(num_towers, int) or num_towers < 1 or num_towers > 10:
            raise ValueError("Number of towers must be between 1 and 10")

        # Remove old tower sheets
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("Tower "):
                wb.remove(wb[sheet_name])

        # Create tower sheets
        for tower in range(1, num_towers + 1):
            logger.info(f"Creating sheet for Tower {tower}")
            tower_sheet = wb.create_sheet(f"Tower {tower}", tower)
            tower_sheet['A1'] = f"Tower {tower} Configuration"
            tower_sheet['A1'].font = Font(bold=True, size=14)
            tower_sheet['A2'] = "Enter details for this tower below."
            headers = [
                "Parameter", "Value", "Notes"
            ]
            for col, header in enumerate(headers, 1):
                cell = tower_sheet.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Tower inputs
            tower_inputs = [
                ("Base Tower Height (in)", "", "Choose from dropdown"),
                ("Lock Controller Location", project_sheet[f'B{tower+4}'].value or "None", "Set in Project Configurator"),
                ("Control Housing", "No", "Yes adds 18\" height"),
                ("Material", "", "M=MDF, P=Polywood, G=Polygood"),
                ("Lock Type", "", "B=Basic, M=Medium, P=Premier"),
                ("Hinge Type", "", "S=Standard, R=Reinforced"),
                ("Third Party Hinges", "No", "Yes = no hinge cost"),
                ("Third Party Locks", "No", "Yes = no lock cost"),
                ("Third Party Material", "No", "Yes = no material cost"),
                ("Third Party Labor", "No", "Yes = no labor cost"),
                ("Third Party CNC", "No", "Yes = no CNC cost")
            ]
            for row, (param, value, note) in enumerate(tower_inputs, 5):
                tower_sheet[f'A{row}'] = param
                tower_sheet[f'B{row}'] = value
                tower_sheet[f'C{row}'] = note

            # Locker configs
            tower_sheet['A16'] = "Locker Configurations"
            tower_sheet['A16'].font = Font(bold=True)
            locker_headers = ["Config Number", "Locker Width (in)", "Number of Lockers", "Locker Heights (in)"]
            for col, header in enumerate(locker_headers, 1):
                cell = tower_sheet.cell(row=17, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            for row in range(18, 23):  # 5 configs
                tower_sheet.cell(row=row, column=1).value = row - 17

            # Data validations
            dv_height = DataValidation(type="list", formula1="Reference!C2", allow_blank=True)
            dv_yes_no = DataValidation(type="list", formula1="Reference!C13", allow_blank=True)
            dv_material = DataValidation(type="list", formula1="Reference!C5", allow_blank=True)
            dv_lock = DataValidation(type="list", formula1="Reference!C7", allow_blank=True)
            dv_hinge = DataValidation(type="list", formula1="Reference!C9", allow_blank=True)
            dv_width = DataValidation(type="list", formula1="Reference!C12", allow_blank=True)
            dv_locker_height = DataValidation(type="list", formula1="Reference!C11", allow_blank=True)
            tower_sheet.add_data_validation(dv_height)
            tower_sheet.add_data_validation(dv_yes_no)
            tower_sheet.add_data_validation(dv_material)
            tower_sheet.add_data_validation(dv_lock)
            tower_sheet.add_data_validation(dv_hinge)
            tower_sheet.add_data_validation(dv_width)
            tower_sheet.add_data_validation(dv_locker_height)
            tower_sheet.add_data_validation(dv_height).add("B5")
            tower_sheet.add_data_validation(dv_yes_no).add("B7")
            tower_sheet.add_data_validation(dv_material).add("B8")
            tower_sheet.add_data_validation(dv_lock).add("B9")
            tower_sheet.add_data_validation(dv_hinge).add("B10")
            for row in range(11, 16):
                tower_sheet.add_data_validation(dv_yes_no).add(f"B{row}")
            for row in range(18, 23):
                tower_sheet.add_data_validation(dv_width).add(f"B{row}")
                tower_sheet.add_data_validation(dv_locker_height).add(f"D{row}")

            tower_sheet.column_dimensions['A'].width = 20
            tower_sheet.column_dimensions['B'].width = 20
            tower_sheet.column_dimensions['C'].width = 30
            tower_sheet.column_dimensions['D'].width = 20

        # Update Summary Sheet
        summary_sheet = wb["Summary"]
        for row in range(5, 15):  # Clear old data
            for col in range(1, 19):
                summary_sheet.cell(row=row, column=col).value = None
        for tower in range(1, num_towers + 1):
            row = tower + 4
            tower_sheet = wb[f"Tower {tower}"]
            summary_sheet.cell(row=row, column=1).value = tower
            summary_sheet.cell(row=row, column=2).value = f"='Tower {tower}'!B5"
            summary_sheet.cell(row=row, column=4).value = f"='Tower {tower}'!B6"
            summary_sheet.cell(row=row, column=5).value = f"='Tower {tower}'!B7"
            for col, cell in [(6, "B8"), (7, "B9"), (8, "B10")]:
                summary_sheet.cell(row=row, column=col).value = f"=IF('Tower {tower}'!{cell}<>\"\",IFERROR(VLOOKUP('Tower {tower}'!{cell},{{{{Reference!C5,Reference!C5;Reference!C7,Reference!C7;Reference!C9,Reference!C9}},1,FALSE),'Tower {tower}'!{cell}),\"\")"
            summary_sheet.cell(row=row, column=3).value = f"=IF(B{row}<>\"\",B{row}+IFERROR(VLOOKUP(D{row},Reference!C20:D20,2,FALSE),0),\"\")"
            summary_sheet.cell(row=row, column=9).value = f"=IF(B{row}<>\"\",TEXTJOIN(\"; \",TRUE,IF('Tower {tower}'!A18:A22<>\"\",IF('Tower {tower}'!B18:B22=\"9\",\"9-inch: \"&'Tower {tower}'!C18:C22&\" lockers, heights: \"&'Tower {tower}'!D18:D22,\"18-inch: \"&'Tower {tower}'!C18:C22&\" lockers, heights: \"&'Tower {tower}'!D18:D22),\"\")),\"\")"
            summary_sheet.cell(row=row, column=10).value = f"=IF(B{row}<>\"\",SUM(IF('Tower {tower}'!A18:A22<>\"\",IF('Tower {tower}'!B18:B22=\"9\",(4*(9*'Tower {tower}'!D18:D22/144+9*'Tower {tower}'!D18:D22/144+9*Reference!B15/144)*2-2*'Tower {tower}'!D18:D22*Reference!B15/144),4*(18*'Tower {tower}'!D18:D22/144+18*'Tower {tower}'!D18:D22/144+18*Reference!B15/144)))),0)+IF(E{row}=\"Yes\",2*(Reference!B14*Reference!B16/144+Reference!B14*Reference!B16/144+Reference!B14*Reference!B15/144),0)+IF(D{row}<>\"\",IF(D{row}=\"Both\",2*(2*(Reference!B14*4/144)+2*(Reference!B14*4/144)+2*(Reference!B14*Reference!B15/144)),IF(D{row}=\"None\",0,2*(Reference!B14*4/144+Reference!B14*4/144+Reference!B14*Reference!B15/144))),0)+2*(Reference!B14*MAX(IF('Tower {tower}'!A18:A22<>\"\",'Tower {tower}'!D18:D22))/144)+2*(Reference!B14*Reference!B15/144),\"\")"
            summary_sheet.cell(row=row, column=11).value = f"=IF(AND(J{row}<>\"\",F{row}<>\"\",'Tower {tower}'!B13<>\"Yes\"),J{row}*IFERROR(VLOOKUP(F{row},Reference!C6:D6,2,FALSE),0),0)"
            summary_sheet.cell(row=row, column=12).value = f"=IF(AND(B{row}<>\"\",'Tower {tower}'!B11<>\"Yes\"),SUM(IF('Tower {tower}'!A18:A22<>\"\",'Tower {tower}'!C18:C22*2))+(IF(E{row}=\"Yes\",2,0))*IFERROR(VLOOKUP(H{row},Reference!C10:D10,2,FALSE),0),0)"
            summary_sheet.cell(row=row, column=13).value = f"=IF(AND(B{row}<>\"\",'Tower {tower}'!B12<>\"Yes\"),SUM(IF('Tower {tower}'!A18:A22<>\"\",'Tower {tower}'!C18:C22))+(IF(E{row}=\"Yes\",1,0))*IFERROR(VLOOKUP(G{row},Reference!C8:D8,2,FALSE),0),0)"
            summary_sheet.cell(row=row, column=14).value = f"=L{row}+M{row}"
            summary_sheet.cell(row=row, column=15).value = f"=IF(AND(B{row}<>\"\",'Tower {tower}'!B14<>\"Yes\"),(SUM(IF('Tower {tower}'!A18:A22<>\"\",'Tower {tower}'!C18:C22*1))+(IF(E{row}=\"Yes\",2,0))+SUM(IF('Tower {tower}'!A18:A22<>\"\",IF('Tower {tower}'!B18:B22=\"9\",'Tower {tower}'!C18:C22*0.5,0))))*Reference!B17,0)"
            summary_sheet.cell(row=row, column=16).value = f"=IF(AND(B{row}<>\"\",'Tower {tower}'!B15<>\"Yes\"),Reference!B18/SUMPRODUCT((B5:B14<>\"\")*1),0)"
            summary_sheet.cell(row=row, column=17).value = f"=SUM(K{row}:P{row})"
            summary_sheet.cell(row=row, column=18).value = f"=Q{row}*Reference!B19"

        # Project totals
        totals_row = num_towers + 5
        summary_sheet.cell(row=totals_row, column=1).value = "Project Totals"
        summary_sheet.cell(row=totals_row, column=1).font = Font(bold=True)
        for col, field in [(10, "J"), (11, "K"), (12, "L"), (13, "M"), (14, "N"), (15, "O"), (16, "P"), (17, "Q"), (18, "R")]:
            summary_sheet.cell(row=totals_row, column=col).value = f"=SUM({field}5:{field}{totals_row-1})"
        summary_sheet.cell(row=totals_row, column=19).value = f"=SUMPRODUCT((B5:B14<>\"\")*Reference!B14)+SUMPRODUCT((B5:B14<>\"\")*(B5:B14<>\"\"))*Reference!B15"

        # Adjust column widths
        for sheet in wb:
            if sheet.title != "Reference":
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    sheet.column_dimensions[column].width = adjusted_width

        # Save the file
        logger.info(f"Saving updated file to {os.path.abspath(filename)}")
        wb.save(filename)
        logger.info(f"Excel file updated successfully: {os.path.abspath(filename)}")

    except Exception as e:
        logger.error(f"Error updating Excel file: {e}")
        raise

if __name__ == "__main__":
    try:
        update_locker_configurator_excel()
    except Exception as e:
        logger.error(f"Script failed: {e}")
    input("Press Enter to exit...")